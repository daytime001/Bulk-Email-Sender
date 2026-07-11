from __future__ import annotations

import json
import os
import re
from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path
from urllib.parse import unquote, urlparse

from bulk_email_sender.models import Recipient

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
EMAIL_HEADERS = {"email", "e-mail", "邮箱", "邮箱地址"}
NAME_HEADERS = {"name", "姓名", "导师姓名", "老师姓名"}
RESEARCH_DIRECTION_HEADERS = {
    "research_direction",
    "research direction",
    "direction",
    "research",
    "研究方向",
    "研究领域",
    "方向",
}
RESEARCH_DIRECTION_KEYS = ("research_direction", "researchDirection", "direction", "research", "研究方向")


class RecipientLoadError(ValueError):
    """Raised when recipient data cannot be parsed safely."""


@dataclass(frozen=True)
class RecipientStats:
    total_rows: int
    valid_rows: int
    sendable_rows: int
    invalid_rows: int
    invalid_email_rows: int
    missing_name_rows: int
    duplicate_rows: int
    empty_rows: int


@dataclass(frozen=True)
class RecipientLoadResult:
    recipients: list[Recipient]
    stats: RecipientStats


def load_recipients(
    file_path: str | Path,
    *,
    raise_on_invalid: bool = True,
) -> RecipientLoadResult:
    path = _normalize_input_path(file_path)
    if not path.exists():
        raise RecipientLoadError(f"Recipient file not found: {path}")

    suffix = path.suffix.lower()
    if suffix == ".json":
        rows = _load_json_rows(path)
    elif suffix in {".xlsx", ".xlsm"}:
        rows = _load_xlsx_rows(path)
    else:
        raise RecipientLoadError(f"Unsupported recipient file format: {suffix}")

    return _normalize_rows(rows, raise_on_invalid=raise_on_invalid)


def _normalize_input_path(file_path: str | Path) -> Path:
    if isinstance(file_path, Path):
        return file_path

    text = os.fsdecode(file_path).strip().strip("\ufeff\u200b\u200c\u200d\u2060")
    if len(text) >= 2 and text[0] == text[-1] and text[0] in {"'", '"'}:
        text = text[1:-1].strip().strip("\ufeff\u200b\u200c\u200d\u2060")

    parsed = urlparse(text)
    if parsed.scheme.lower() == "file":
        path_text = unquote(parsed.path)
        if parsed.netloc and parsed.netloc.lower() != "localhost":
            path_text = f"//{parsed.netloc}{path_text}"
        if re.match(r"^/[A-Za-z]:[\\/]", path_text):
            path_text = path_text[1:]
        return Path(path_text)

    return Path(text)


def _load_json_rows(path: Path) -> list[tuple[int, object, object, object]]:
    payload = json.loads(_read_json_text(path))

    rows: list[tuple[int, object, object, object]] = []
    if isinstance(payload, dict):
        for index, (email, name) in enumerate(payload.items(), start=1):
            rows.append((index, email, name, None))
        return rows

    if isinstance(payload, list):
        for index, item in enumerate(payload, start=1):
            if not isinstance(item, dict):
                raise RecipientLoadError(f"Invalid JSON row at index {index}: expected object")
            rows.append((index, item.get("email"), item.get("name"), _get_research_direction_value(item)))
        return rows

    raise RecipientLoadError("Invalid JSON format: expected object or list")


def _read_json_text(path: Path) -> str:
    errors: list[str] = []
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError as exc:
            errors.append(f"{encoding}: {exc}")
    raise RecipientLoadError("JSON 文件编码无法识别，请另存为 UTF-8 后重试：" + " | ".join(errors))


def _get_research_direction_value(item: dict[str, object]) -> object:
    for key in RESEARCH_DIRECTION_KEYS:
        if key in item:
            return item.get(key)
    return None


def _load_xlsx_rows(path: Path) -> list[tuple[int, object, object, object]]:
    from openpyxl import load_workbook

    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        value_rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not value_rows:
        return []

    first_row = value_rows[0]
    header_map = _detect_header_map(first_row)
    data_rows = value_rows[1:] if header_map else value_rows

    if header_map is None:
        if _looks_like_email(_cell_to_text(first_row[0] if len(first_row) > 0 else None)):
            email_idx, name_idx, research_direction_idx = 0, 1, 2
        else:
            raise RecipientLoadError(
                "Unable to detect XLSX columns. Use headers '邮箱/姓名' "
                "or place email in column A and name in column B."
            )
    else:
        email_idx, name_idx, research_direction_idx = header_map

    rows: list[tuple[int, object, object, object]] = []
    for row_number, row in enumerate(data_rows, start=2 if header_map else 1):
        email = row[email_idx] if len(row) > email_idx else None
        name = row[name_idx] if len(row) > name_idx else None
        research_direction = (
            row[research_direction_idx]
            if research_direction_idx is not None and len(row) > research_direction_idx
            else None
        )
        rows.append((row_number, email, name, research_direction))
    return rows


def _detect_header_map(row: Iterable[object]) -> tuple[int, int, int | None] | None:
    normalized = [_cell_to_text(value).strip().lower() for value in row]
    email_idx = None
    name_idx = None
    research_direction_idx = None
    for idx, value in enumerate(normalized):
        if value in EMAIL_HEADERS and email_idx is None:
            email_idx = idx
        if value in NAME_HEADERS and name_idx is None:
            name_idx = idx
        if value in RESEARCH_DIRECTION_HEADERS and research_direction_idx is None:
            research_direction_idx = idx
    if email_idx is None or name_idx is None:
        return None
    return email_idx, name_idx, research_direction_idx


def _normalize_rows(
    rows: list[tuple[int, object, object, object]],
    *,
    raise_on_invalid: bool,
) -> RecipientLoadResult:
    seen: set[str] = set()
    recipients: list[Recipient] = []
    invalid_messages: list[str] = []
    sendable_rows = 0
    invalid_email_rows = 0
    missing_name_rows = 0
    duplicate_rows = 0
    empty_rows = 0

    for row_number, raw_email, raw_name, raw_research_direction in rows:
        email = _cell_to_text(raw_email).strip()
        name = _cell_to_text(raw_name).strip()
        research_direction = _cell_to_text(raw_research_direction).strip()

        if not email and not name and not research_direction:
            empty_rows += 1
            continue

        if not _looks_like_email(email):
            invalid_email_rows += 1
            invalid_messages.append(f"row {row_number}: invalid email '{email}'")
            continue

        if not name:
            missing_name_rows += 1
            invalid_messages.append(f"row {row_number}: missing recipient name")
            continue

        sendable_rows += 1
        email_key = email.lower()
        if email_key in seen:
            duplicate_rows += 1
            continue

        seen.add(email_key)
        recipients.append(Recipient(email=email, name=name, research_direction=research_direction))

    if raise_on_invalid and invalid_messages:
        details = "; ".join(invalid_messages[:20])
        raise RecipientLoadError(f"Recipient file contains invalid rows: {details}")

    total_rows = len(rows)
    invalid_rows = invalid_email_rows + missing_name_rows
    stats = RecipientStats(
        total_rows=total_rows,
        valid_rows=len(recipients),
        sendable_rows=sendable_rows,
        invalid_rows=invalid_rows,
        invalid_email_rows=invalid_email_rows,
        missing_name_rows=missing_name_rows,
        duplicate_rows=duplicate_rows,
        empty_rows=empty_rows,
    )
    return RecipientLoadResult(recipients=recipients, stats=stats)


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def _looks_like_email(value: str) -> bool:
    return bool(EMAIL_RE.match(value))
